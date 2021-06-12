import os, time
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter





column_widths = []
dateTimeObj = datetime.now()
dateStampStr = dateTimeObj.strftime("%d%b%Y")
timeStampStr = dateTimeObj.strftime("%H:%M:%S")
#print (dateStampStr,timeStampStr)

#filepath = "Data_"+dateStampStr+"_"+dateTimeObj.strftime("%H%M%S")

ValidBarcodeNumbers = ['01-STK', '02-STK', '03-STK', '04-STK', '05-STK'] #for hardware data


"""


wb = openpyxl.Workbook()
ws = wb.active
column_letters = ['A','B','C','D','E','F']
for column_letter in column_letters:
    ws.column_dimensions[column_letter].bestFit = True
x = 1

ws.cell(row=1, column=1, value='Barcode')
ws.cell(row= 1, column=2, value='Name of hardware')
ws.cell(row=1, column=3, value='Employee')

ws.cell(row=1, column=4, value='Borrow')
ws.cell(row=1, column=5, value='Return')
ws.cell(row=1, column=6, value='Remarks')
"""

colCounter=1 # column pointer


wb = load_workbook(filename='Data.xlsx')
ws = wb['Sheet1']

for row in range(2,ws.max_row):#check if cells in barcode column is empty
    if(ws.cell(row,1).value is not None):
            print(ws.cell(row,1).value)
            colCounter=colCounter+1
            print (colCounter)


while True:
    barcodeNumber = input("Scan barcode: ")

    if barcodeNumber in ValidBarcodeNumbers:
        BorrowOrReturn = input("Borrow or Return: ")  # Borrow or Return
        if (BorrowOrReturn == 'Borrow' or BorrowOrReturn == 'B'): #Borrow
            dateTimeObj = datetime.now()
            dateStampStr = dateTimeObj.strftime("%d-%b-%Y")
            timeStampStr = dateTimeObj.strftime("%H:%M:%S")
            ws.cell(row=colCounter + 1, column=1, value=str(barcodeNumber))
            ws.cell(row=colCounter + 1, column=4, value=str(str(dateStampStr) + ', ' + str(timeStampStr)))
            # ws.cell(row=colCounter + 1, column=6, value=str(Remarks))
            print('Borrowed ' + barcodeNumber + ' on ' + dateStampStr + ' ' + timeStampStr)
            wb.save(filename='Data.xlsx')
            colCounter = colCounter + 1

        elif (BorrowOrReturn == 'Return' or BorrowOrReturn == 'R'):
            for cell in ws['A']:
                if (cell.value is not None):
                    if barcodeNumber in cell.value:
                        print(
                            'Found header with name: {} at row: {} and column: {}. In cell {}'.format(cell.value,
                                                                                                      cell.row,
                                                                                                      cell.column,
                                                                                                      cell))
                        retCellrow = cell.row
            dateTimeObj = datetime.now()
            dateStampStr = dateTimeObj.strftime("%d-%b-%Y")
            timeStampStr = dateTimeObj.strftime("%H:%M:%S")
            ws.cell(row=retCellrow, column=1, value=str(barcodeNumber))
            ws.cell(row=retCellrow, column=5, value=str(str(dateStampStr) + ', ' + str(timeStampStr)))
            ws.cell(row=retCellrow, column=6, value=str(Remarks))
            print('Time: ', timeStampStr)

            Remarks = input("Remarks: ")
            wb.save(filename='Data.xlsx')
            colCounter = colCounter + 1

        elif (BorrowOrReturn == 'Check' or BorrowOrReturn == 'C'): #Board availability
            for cell in ws['A']:
                if (cell.value is not None):
                    if barcodeNumber in cell.value:
                        #print('Found header with name: {} at row: {} and column: {}. In cell {}'.format(cell.value,cell.row,cell.column,cell))
                        retCellrow = cell.row
                        print (barcodeNumber+' is not available')
                        print('Currently with ' + " " + ws.cell(retCellrow, 3).value)
                        break
                else:
                    print (barcodeNumber+ ' is Available')
                    break
    else:
        print("INVALID BARCODE DATA!")











"""
    BorrowOrReturn = input("Borrow or Return: ") #Borrow or Return
    #Remarks = input ("Remarks: ")
    if barcodeNumber in ValidBarcodeNumbers and (BorrowOrReturn == 'Borrow' or BorrowOrReturn == 'B'):
        dateTimeObj = datetime.now()
        dateStampStr = dateTimeObj.strftime("%d-%b-%Y")
        timeStampStr = dateTimeObj.strftime("%H:%M:%S")
        ws.cell(row=colCounter + 1, column=1, value=str(barcodeNumber))
        ws.cell(row=colCounter + 1, column=4, value=str(str(dateStampStr) + ', ' + str(timeStampStr)))
        #ws.cell(row=colCounter + 1, column=6, value=str(Remarks))
        print('Borrowed '+barcodeNumber+' on ' + dateStampStr + ' ' + timeStampStr)

        wb.save(filename='Data.xlsx')

        colCounter=colCounter+1

    elif barcodeNumber in ValidBarcodeNumbers and (BorrowOrReturn == 'Return' or BorrowOrReturn == 'R'):
        Remarks = input("Remarks: ")

        #FIND borrowed board uisng barcode scan
        
        for cell in ws['A']:
            if (cell.value is not None):  # We need to check that the cell is not empty.
                if barcodeNumber in cell.value:  # Check if the value of the cell contains the text 'Table'
                    print(
                        'Found header with name: {} at row: {} and column: {}. In cell {}'.format(cell.value, cell.row,cell.column, cell))
                    retCellrow = cell.row

        dateTimeObj = datetime.now()
        dateStampStr = dateTimeObj.strftime("%d-%b-%Y")
        timeStampStr = dateTimeObj.strftime("%H:%M:%S")
        ws.cell(row=retCellrow, column=1, value=str(barcodeNumber))
        ws.cell(row=retCellrow, column=5, value=str(str(dateStampStr) + ', ' +  str(timeStampStr)))
        ws.cell(row=retCellrow, column=6, value=str(Remarks))
        print('Time: ', timeStampStr)


        wb.save(filename='Data.xlsx')

        colCounter=colCounter+1

    elif BorrowOrReturn == 'Quit' or BorrowOrReturn == 'quit':
        exit()



    else:
        print ("INVALID BARCODE DATA!")
"""


